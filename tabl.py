import os
import openpyxl
import pandas as pd

wb = openpyxl.Workbook()
sheet = wb.active

# 	СОЗДАЮ КОЛОНКИ В ТАБЛИЦЕ
c1 = sheet.cell(row = 1, column = 1)
c1.value = "time"
c1 = sheet.cell(row = 1, column = 2)
c1.value = "Y1"
c1 = sheet.cell(row = 1, column = 3)
c1.value = "dist1"
c1 = sheet.cell(row = 1, column = 4)
c1.value = "Y2"
c1 = sheet.cell(row = 1, column = 5)
c1.value = "dist2"
c1 = sheet.cell(row = 1, column = 8)
c1.value = "dist_betv12"
c1 = sheet.cell(row = 1, column = 6)
c1.value = "Y3"
c1 = sheet.cell(row = 1, column = 7)
c1.value = "dist3"
c1 = sheet.cell(row = 1, column = 9)
c1.value = "dist_betv23"
c1 = sheet.cell(row = 1, column = 10)
c1.value = "dist_betv13"

c1 = sheet.cell(row = 1, column = 13)
c1.value = "Y"
c1 = sheet.cell(row = 1, column = 12)
c1.value = "X"
